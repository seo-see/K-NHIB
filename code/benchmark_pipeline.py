"""
K-NHIB - Korean National Health Insurance Benchmark - LLM Evaluation Pipeline
6 models (Anthropic, Google, OpenAI) × guideline-input conditions × 222 cases

Models (3 providers × 2): Claude (Opus / Sonnet), GPT (5.4 / mini), Gemini (Pro / Flash)
Conditions (guideline material supplied to the model):
  - pdf            : guideline PDF attached natively   (main analysis)
  - text_md        : guideline as plain text extracted from the PDF (pdfplumber)  (sensitivity)
  - pdf_websearch  : guideline PDF + web search          (sensitivity)
  - pdf_structure  : structure-guided prompt             (sensitivity)
"""

import os
import time
import json
import csv
import base64
import random
import unicodedata
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
import pandas as pd


# ──────────────────────────────────────────────
# 0. Configuration
# ──────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent          # benchmark/code
REPO_DIR = BASE_DIR.parent                            # benchmark/
DATA_DIR = REPO_DIR / "data"
BENCHMARK_DIR = DATA_DIR / "benchmark"
ASSETS_DIR = DATA_DIR / "guidelines"                  # find_asset() searches guideline PDFs here


def load_local_env() -> None:
    """Simple .env loader. Does not overwrite existing environment variables."""
    for candidate in (BASE_DIR / ".env", BASE_DIR / ".env.local"):
        if not candidate.exists():
            continue
        for raw_line in candidate.read_text().splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip("'").strip('"')
            os.environ.setdefault(key, value)


def _normalize_text(text: str) -> str:
    return unicodedata.normalize("NFC", text or "")


def find_asset(filename: str = "", *, contains: str = "", suffix: str = "") -> Path:
    """Safely search for a filename in the guidelines directory."""
    normalized_filename = _normalize_text(filename)
    normalized_contains = _normalize_text(contains)

    for path in ASSETS_DIR.iterdir():
        normalized_name = _normalize_text(path.name)
        if filename and normalized_name == normalized_filename:
            return path
        if contains and normalized_contains in normalized_name and (
            not suffix or normalized_name.endswith(suffix)
        ):
            return path

    target = filename or contains
    raise FileNotFoundError(f"Asset not found: {target}")


EXCEL_PATH = BENCHMARK_DIR / "K-NHIB_GY.xlsx"

# Per-cancer guideline PDF paths
PDF_PATHS = {
    "자궁경부암": find_asset(contains="Cervical", suffix=".pdf"),
    "자궁내막암": find_asset(contains="Uterine", suffix=".pdf"),
    "난소암": find_asset(contains="Ovarian", suffix=".pdf"),
}

# Cancer mapping (sheet name -> cancer_type key)
# Sheet names are English (cervical/uterine/ovarian); cancer_type values stay Korean.
# Korean values keep PDF_PATHS lookup and the prompt "Cancer type: ..." input identical to the study.
FORWARD_SHEET_MAP = {
    "cervical": "자궁경부암",
    "uterine": "자궁내막암",
    "ovarian": "난소암",
}
# API keys (managed via environment variables)
load_local_env()
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY    = os.environ.get("OPENAI_API_KEY", "")
GEMINI_API_KEY    = os.environ.get("GEMINI_API_KEY", "")

# Model codes
SUPPORTED_MODEL_CODES = {
    "claude": [
        "claude-opus-4-6",
        "claude-sonnet-4-6",
    ],
    "gpt": [
        "gpt-5.4",
        "gpt-5-mini-2025-08-07",
    ],
    "gemini": [
        "gemini-3.1-pro-preview",
        "gemini-3-flash-preview",
    ],
}
MODEL_ALIASES = {
    "claude": "claude-opus-4-6",
    "claude-fast": "claude-sonnet-4-6",
    "gpt": "gpt-5.4",
    "gpt-fast": "gpt-5-mini-2025-08-07",
    "gemini": "gemini-3.1-pro-preview",
    "gemini-fast": "gemini-3-flash-preview",
}
MODEL_PROVIDER_MAP = {
    model_code: provider
    for provider, model_codes in SUPPORTED_MODEL_CODES.items()
    for model_code in model_codes
}
MODEL_PROVIDER_MAP.update({
    alias: MODEL_PROVIDER_MAP[model_code]
    for alias, model_code in MODEL_ALIASES.items()
})
SUPPORTED_MODEL_NAMES = list(MODEL_ALIASES.keys()) + [
    model_code
    for model_codes in SUPPORTED_MODEL_CODES.values()
    for model_code in model_codes
]

OUTPUT_DIR = REPO_DIR / "results"
OUTPUT_DIR.mkdir(exist_ok=True)

API_MAX_RETRIES = max(0, int(os.environ.get("API_MAX_RETRIES", "2")))
API_RETRY_BASE_SECONDS = float(os.environ.get("API_RETRY_BASE_SECONDS", "1.0"))
API_RETRY_MAX_SECONDS = float(os.environ.get("API_RETRY_MAX_SECONDS", "8.0"))
API_RETRY_JITTER_SECONDS = float(os.environ.get("API_RETRY_JITTER_SECONDS", "0.25"))
RETRYABLE_STATUS_CODES = {408, 409, 429, 500, 502, 503, 504}
GEMINI_HTTP_TIMEOUT_MS = int(os.environ.get("GEMINI_HTTP_TIMEOUT_MS", "300000"))
GENERATION_SETTINGS_POLICY = "provider_defaults"
RESULT_KEY_COLUMNS = ["id", "model", "condition"]
RESULT_COLUMNS = [
    "id",
    "cancer_type",
    "class",
    "regimen_code",
    "regimen",
    "expected",
    "model",
    "model_provider",
    "model_version",
    "generation_settings_policy",
    "condition",
    "run_id",
    "predicted",
    "reason",
    "correct",
    "raw_response",
    "search_queries",
    "latency",
    "total_latency",
    "retry_sleep_seconds",
    "attempt_count",
    "input_tokens",
    "output_tokens",
    "timestamp",
]


# ──────────────────────────────────────────────
# 1. Data loading
# ──────────────────────────────────────────────

def load_benchmark(excel_path: str | Path) -> pd.DataFrame:
    """Load forward-direction cases into a single DataFrame."""
    wb = openpyxl.load_workbook(excel_path)
    rows = []

    for sheet_name, cancer_type in FORWARD_SHEET_MAP.items():
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            if row.get("ID"):
                row["cancer_type"] = cancer_type
                row["regimen_code"] = row.get("regimen_code", "")
                row["regimen"] = row.get("regimen", "")
                row["expected_label"] = row.get("expected", "")
                rows.append(row)

    df = pd.DataFrame(rows)
    print(f"Loaded {len(df)} cases")
    print(df.groupby("cancer_type").size())
    return df


def load_saved_results(path: str | Path) -> pd.DataFrame:
    result_path = Path(path)
    if not result_path.exists():
        raise FileNotFoundError(f"Previous results file not found: {result_path}")

    suffix = result_path.suffix.lower()
    if suffix == ".xlsx":
        return pd.read_excel(result_path, sheet_name="raw")
    if suffix == ".csv":
        return pd.read_csv(result_path)

    raise ValueError("Unsupported results file format. Use csv or xlsx.")


def build_retry_task_plan(
    benchmark_df: pd.DataFrame,
    previous_results: pd.DataFrame,
    models: list[str],
    conditions: list[str],
) -> pd.DataFrame:
    required_columns = {"id", "model", "condition", "predicted"}
    missing_columns = required_columns - set(previous_results.columns)
    if missing_columns:
        raise ValueError(
            "Previous results file is missing required columns: "
            + ", ".join(sorted(missing_columns))
        )

    failed = previous_results[previous_results["predicted"].astype(str) == "error"].copy()
    failed = failed[failed["model"].isin(models) & failed["condition"].isin(conditions)]
    failed = failed[RESULT_KEY_COLUMNS].drop_duplicates(subset=RESULT_KEY_COLUMNS)

    benchmark_rows = benchmark_df.copy()
    benchmark_rows["id"] = benchmark_rows["ID"]
    task_plan = failed.merge(benchmark_rows, on="id", how="inner")
    task_plan["__model"] = task_plan["model"]
    task_plan["__condition"] = task_plan["condition"]
    return task_plan


def merge_results_frames(previous_results: pd.DataFrame, rerun_results: pd.DataFrame) -> pd.DataFrame:
    merged = pd.concat([previous_results, rerun_results], ignore_index=True)
    merged = merged.drop_duplicates(subset=RESULT_KEY_COLUMNS, keep="last")
    sort_columns = [
        column
        for column in ["cancer_type", "id", "condition", "model"]
        if column in merged.columns
    ]
    if sort_columns:
        merged = merged.sort_values(sort_columns).reset_index(drop=True)
    return merged


def append_result_csv_row(csv_path: str | Path, row: dict) -> None:
    """Persist results immediately during long runs."""
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not csv_path.exists() or csv_path.stat().st_size == 0
    normalized_row = {column: row.get(column, "") for column in RESULT_COLUMNS}

    with csv_path.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=RESULT_COLUMNS)
        if write_header:
            writer.writeheader()
        writer.writerow(normalized_row)


# ──────────────────────────────────────────────
# 2. PDF loading
# ──────────────────────────────────────────────

def load_pdf_bytes(pdf_path: str | Path) -> bytes:
    with open(pdf_path, "rb") as f:
        return f.read()


def load_pdf_base64(pdf_path: str | Path) -> str:
    return base64.standard_b64encode(load_pdf_bytes(pdf_path)).decode("utf-8")


def load_pdf_text(pdf_path: str | Path) -> str:
    """Extract text from PDF (used by the text_md condition).

    Delegates to the pluggable parser in pdf_parser.py so the parsing backend can be
    swapped for experiments without touching the pipeline. The study used pdfplumber.
    """
    from pdf_parser import extract_text
    return extract_text(pdf_path)


# PDF cache
_pdf_text_cache: dict = {}
_pdf_base64_cache: dict = {}
_pdf_bytes_cache: dict = {}


def get_pdf_text(pdf_path: str | Path) -> str:
    pdf_path = str(pdf_path)
    if pdf_path not in _pdf_text_cache:
        text = load_pdf_text(pdf_path)
        _pdf_text_cache[pdf_path] = text
        # Persist the parsed text as a reproducible artifact (under results/, gitignored).
        # Source guideline PDFs ship with the repo; parsed text is generated on run.
        parsed_dir = OUTPUT_DIR / "parsed_guidelines"
        parsed_dir.mkdir(parents=True, exist_ok=True)
        (parsed_dir / f"{Path(pdf_path).stem}.txt").write_text(text, encoding="utf-8")
    return _pdf_text_cache[pdf_path]


def get_pdf_base64(pdf_path: str | Path) -> str:
    pdf_path = str(pdf_path)
    if pdf_path not in _pdf_base64_cache:
        _pdf_base64_cache[pdf_path] = load_pdf_base64(pdf_path)
    return _pdf_base64_cache[pdf_path]


def get_pdf_bytes(pdf_path: str | Path) -> bytes:
    pdf_path = str(pdf_path)
    if pdf_path not in _pdf_bytes_cache:
        _pdf_bytes_cache[pdf_path] = load_pdf_bytes(pdf_path)
    return _pdf_bytes_cache[pdf_path]


def require_api_key(provider: str, api_key: str) -> None:
    if not api_key:
        raise RuntimeError(
            f"{provider} API key is missing. Set the {provider}_API_KEY environment variable."
        )


class APIRequestError(RuntimeError):
    def __init__(self, message: str, *, status_code: int | None = None, retryable: bool | None = None):
        super().__init__(message)
        self.status_code = status_code
        self.retryable = retryable


def get_exception_status_code(exc: Exception) -> int | None:
    candidates = [
        getattr(exc, "status_code", None),
        getattr(getattr(exc, "response", None), "status_code", None),
        getattr(getattr(exc, "response", None), "status", None),
    ]
    for candidate in candidates:
        if candidate is None:
            continue
        try:
            return int(candidate)
        except (TypeError, ValueError):
            continue
    return None


def is_retryable_exception(exc: Exception) -> bool:
    if isinstance(exc, APIRequestError) and exc.retryable is not None:
        return exc.retryable

    if isinstance(exc, (httpx.TimeoutException, httpx.NetworkError, httpx.TransportError)):
        return True

    status_code = get_exception_status_code(exc)
    if status_code in RETRYABLE_STATUS_CODES:
        return True

    try:
        import anthropic

        retryable_types = tuple(
            exc_type
            for exc_type in (
                getattr(anthropic, "APIConnectionError", None),
                getattr(anthropic, "APITimeoutError", None),
                getattr(anthropic, "RateLimitError", None),
                getattr(anthropic, "InternalServerError", None),
            )
            if exc_type is not None
        )
        if retryable_types and isinstance(exc, retryable_types):
            return True
    except Exception:
        pass

    try:
        from google.api_core import exceptions as google_exceptions

        retryable_google_types = (
            google_exceptions.TooManyRequests,
            google_exceptions.ResourceExhausted,
            google_exceptions.ServiceUnavailable,
            google_exceptions.DeadlineExceeded,
            google_exceptions.InternalServerError,
            google_exceptions.BadGateway,
            google_exceptions.GatewayTimeout,
        )
        if isinstance(exc, retryable_google_types):
            return True
    except Exception:
        pass

    message = str(exc).lower()
    return any(
        token in message
        for token in (
            "rate limit",
            "too many requests",
            "temporarily unavailable",
            "timeout",
            "timed out",
            "connection reset",
            "connection error",
            "server error",
            "service unavailable",
        )
    )


def with_retries(label: str, func):
    total_attempts = API_MAX_RETRIES + 1
    total_elapsed = 0.0
    retry_sleep_total = 0.0

    for attempt in range(1, total_attempts + 1):
        attempt_start = time.time()
        try:
            result = func()
            api_latency = time.time() - attempt_start
            total_elapsed += api_latency
            return result, {
                "attempt_count": attempt,
                "api_latency": round(api_latency, 3),
                "total_latency": round(total_elapsed, 3),
                "retry_sleep_seconds": round(retry_sleep_total, 3),
            }
        except Exception as exc:
            api_latency = time.time() - attempt_start
            total_elapsed += api_latency
            setattr(
                exc,
                "_retry_metrics",
                {
                    "attempt_count": attempt,
                    "api_latency": round(api_latency, 3),
                    "total_latency": round(total_elapsed, 3),
                    "retry_sleep_seconds": round(retry_sleep_total, 3),
                },
            )
            if attempt >= total_attempts or not is_retryable_exception(exc):
                raise

            backoff = min(
                API_RETRY_BASE_SECONDS * (2 ** (attempt - 1)),
                API_RETRY_MAX_SECONDS,
            )
            jitter = random.uniform(0, API_RETRY_JITTER_SECONDS)
            wait_seconds = backoff + jitter
            print(
                f"RETRY {label} [{attempt}/{total_attempts - 1}] "
                f"after {wait_seconds:.2f}s due to: {exc}"
            )
            time.sleep(wait_seconds)
            retry_sleep_total += wait_seconds
            total_elapsed += wait_seconds


def extract_openai_response_text(payload: dict) -> str:
    if isinstance(payload.get("output_text"), str) and payload["output_text"].strip():
        return payload["output_text"].strip()

    texts = []
    for item in payload.get("output", []):
        if item.get("type") != "message":
            continue
        for content in item.get("content", []):
            if content.get("type") in {"output_text", "text"}:
                text = content.get("text", "")
                if text:
                    texts.append(text)
    return "\n".join(texts).strip()


def extract_openai_search_queries(payload: dict) -> list[str]:
    """Extract web_search_call queries from the Responses API output"""
    queries = []
    for item in payload.get("output", []):
        if item.get("type") == "web_search_call":
            q = item.get("query", "") or (item.get("action", {}) or {}).get("query", "")
            if q:
                queries.append(q)
    return queries


def extract_anthropic_text(content_blocks) -> str:
    texts = []
    for block in content_blocks:
        if getattr(block, "type", None) == "text":
            text = getattr(block, "text", "")
            if text:
                texts.append(text.strip())
    return "\n".join(texts).strip()


def call_openai_responses_api(payload: dict) -> dict:
    require_api_key("OPENAI", OPENAI_API_KEY)

    try:
        with httpx.Client(timeout=120.0) as client:
            response = client.post(
                "https://api.openai.com/v1/responses",
                headers={
                    "Authorization": f"Bearer {OPENAI_API_KEY}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )
    except httpx.HTTPError as exc:
        raise APIRequestError(f"OpenAI API connection error: {exc}", retryable=True) from exc

    if response.is_error:
        detail = response.text.strip()
        raise APIRequestError(
            f"OpenAI API error ({response.status_code}): {detail}",
            status_code=response.status_code,
            retryable=response.status_code in RETRYABLE_STATUS_CODES,
        )

    data = response.json()
    usage = data.get("usage", {})

    return {
        "raw": extract_openai_response_text(data),
        "search_queries": extract_openai_search_queries(data),
        "input_tokens": usage.get("input_tokens", 0),
        "output_tokens": usage.get("output_tokens", 0),
        "model_version": data.get("model", payload.get("model", "")),
    }


def resolve_model_name(model_name: str, condition: str | None = None) -> str:
    del condition
    if model_name in MODEL_ALIASES:
        return MODEL_ALIASES[model_name]
    if model_name in MODEL_PROVIDER_MAP:
        return model_name
    raise KeyError(f"Unsupported model: {model_name}")


def resolve_model_provider(model_name: str) -> str:
    if model_name in MODEL_PROVIDER_MAP:
        return MODEL_PROVIDER_MAP[model_name]
    raise KeyError(f"Unsupported model: {model_name}")


def validate_runtime(selected_models: list[str], dry_run: bool) -> None:
    if dry_run:
        return

    required_keys = {
        "claude": ("ANTHROPIC_API_KEY", ANTHROPIC_API_KEY),
        "gpt": ("OPENAI_API_KEY", OPENAI_API_KEY),
        "gemini": ("GEMINI_API_KEY", GEMINI_API_KEY),
    }
    missing = [
        env_name
        for model in selected_models
        for provider in [resolve_model_provider(model)]
        for env_name, env_value in [required_keys[provider]]
        if not env_value
    ]
    if missing:
        formatted = ", ".join(sorted(set(missing)))
        raise SystemExit(f"Missing API keys required to run: {formatted}")


# ──────────────────────────────────────────────
# 3. Prompt builders
# ──────────────────────────────────────────────

FORWARD_SYSTEM_PROMPT = """You are an expert reviewer of Korean National Health Insurance reimbursement for oncology regimens.
Given the clinical and administrative attributes and any provided material, decide whether the named regimen is reimbursable.

Respond with exactly one JSON object using this schema:
{"decision":"eligible|ineligible|undeterminable","reason":"<one short sentence>"}

Outcome definitions:
- eligible: All required conditions are met.
- ineligible: 1 or more condition is explicitly not met, regardless of whether other conditions are unevaluable. The condition is present but violates the criterion.
- undeterminable: No condition is explicitly not met, but 1 or more condition is unevaluable due to absent clinical information. Because no criterion is violated, the case cannot be classified as ineligible, yet a determination cannot be reached.
"""

# System prompt for the structure-guided sensitivity condition (`pdf_structure`):
# injects the guideline hierarchy (L1-L4) and asks for step-by-step per-condition verification.
STRUCTURE_GUIDED_SYSTEM_PROMPT = """You are an expert reviewer of Korean National Health Insurance reimbursement for oncology regimens.
Given the clinical and administrative attributes and the provided guideline document, decide whether the named regimen is reimbursable.

Respond with exactly one JSON object using this schema:
{"decision":"eligible|ineligible|undeterminable","reason":"<one short sentence>"}

Outcome definitions:
- eligible: All required conditions are met.
- ineligible: 1 or more condition is explicitly not met, regardless of whether other conditions are unevaluable. The condition is present but violates the criterion.
- undeterminable: No condition is explicitly not met, but 1 or more condition is unevaluable due to absent clinical information. Because no criterion is violated, the case cannot be classified as ineligible, yet a determination cannot be reached.

The reimbursement guideline is organized hierarchically:
- [L1] Global caveat: cancer-type-level conditions applied to all regimens (e.g., default scope of covered indications).
- [L2] Treatment intent: regimens grouped by treatment intent.
- [L3] Regimen [Indication]: the specific regimen and its indication.
- [L4] Footnote: externalized conditions at the document tail, applied via explicit cross-reference (e.g., "see Note 1").

Verify each eligibility condition step by step against its corresponding level:
1. Check the [L1] global caveat for the cancer type.
2. Identify the [L2] treatment intent and the matching [L3] regimen and indication.
3. Apply any [L4] footnote conditions cross-referenced by the regimen.
For each condition, judge whether it is met, not met, or unevaluable from the provided attributes, then aggregate these to the case-level outcome.
"""

MAX_RESPONSE_TOKENS = int(os.environ.get("MAX_RESPONSE_TOKENS", "8192"))

def build_forward_user_prompt(row: dict) -> str:
    return f"""Cancer type: {row['cancer_type']}
Regimen: {row['regimen']}
Clinical and administrative attributes: {row['attributes']}

Task: Determine whether this patient is eligible for reimbursement for the regimen above."""


def get_system_prompt(row: dict, condition: str | None = None) -> str:
    if condition == "pdf_structure":
        return STRUCTURE_GUIDED_SYSTEM_PROMPT
    return FORWARD_SYSTEM_PROMPT


def build_user_prompt(row: dict) -> str:
    return build_forward_user_prompt(row)


def extract_json_object(raw: str) -> dict | None:
    text = raw.strip()
    candidates = [text]

    if text.startswith("```"):
        stripped_lines = [
            line for line in text.splitlines()
            if not line.strip().startswith("```")
        ]
        candidates.append("\n".join(stripped_lines).strip())

    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and start < end:
        candidates.append(text[start:end + 1].strip())

    for candidate in candidates:
        if not candidate:
            continue
        try:
            parsed = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(parsed, dict):
            return parsed
    return None


def parse_model_output(raw: str) -> dict:
    parsed_json = extract_json_object(raw)
    reason = ""

    if parsed_json:
        decision_source = str(
            parsed_json.get("decision")
            or parsed_json.get("answer")
            or parsed_json.get("result")
            or parsed_json.get("eligibility")
            or ""
        ).strip()
        reason = str(
            parsed_json.get("reason")
            or parsed_json.get("rationale")
            or parsed_json.get("explanation")
            or ""
        ).strip()
        decision = parse_answer(decision_source)
        if decision != "invalid":
            return {
                "decision": decision,
                "reason": reason,
                "normalized_raw": json.dumps(parsed_json, ensure_ascii=False),
            }

    return {
        "decision": parse_answer(raw),
        "reason": reason,
        "normalized_raw": raw,
    }


# ──────────────────────────────────────────────
# 4. Per-model call functions
# ──────────────────────────────────────────────

def call_claude(row: dict, condition: str, pdf_path: str | Path = None, model_name: str = "claude") -> dict:
    import anthropic
    require_api_key("ANTHROPIC", ANTHROPIC_API_KEY)
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    content = []
    uses_pdf_native = condition in {"pdf", "pdf_websearch", "pdf_structure"}
    uses_text = condition in {"text_md", "text_md_websearch"}
    uses_websearch = condition in {"websearch", "pdf_websearch", "text_md_websearch"}

    # Condition: native PDF attachment
    if uses_pdf_native and pdf_path:
        content.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": get_pdf_base64(pdf_path),
            },
            "cache_control": {"type": "ephemeral"},  # cost saving
        })
    # Condition: extracted-text attachment (for sensitivity analysis)
    elif uses_text and pdf_path:
        content.append({
            "type": "text",
            "text": f"[급여 고시 원문]\n{get_pdf_text(pdf_path)}",
        })

    content.append({"type": "text", "text": build_user_prompt(row)})

    # Condition: Web Search
    tools = []
    if uses_websearch:
        tools = [{"type": "web_search_20250305", "name": "web_search"}]

    kwargs = {
        "model": resolve_model_name(model_name, condition),
        "max_tokens": MAX_RESPONSE_TOKENS,
        "system": get_system_prompt(row, condition),
        "messages": [{"role": "user", "content": content}],
    }
    if tools:
        kwargs["tools"] = tools

    response, retry_metrics = with_retries(
        f"{resolve_model_name(model_name)}|{condition}",
        lambda: client.messages.create(**kwargs),
    )

    # With web search, intermediate reasoning/tool text may be interleaved, so concatenate all text.
    answer = extract_anthropic_text(response.content)

    # Extract search queries (websearch condition)
    search_queries: list[str] = []
    for block in response.content:
        if getattr(block, "type", None) == "tool_use" and getattr(block, "name", "") == "web_search":
            q = getattr(block, "input", {}).get("query", "")
            if q:
                search_queries.append(q)

    parsed_output = parse_model_output(answer)

    return {
        "raw": parsed_output["normalized_raw"],
        "parsed": parsed_output["decision"],
        "reason": parsed_output["reason"],
        "search_queries": " | ".join(search_queries),
        "latency": retry_metrics["api_latency"],
        "total_latency": retry_metrics["total_latency"],
        "retry_sleep_seconds": retry_metrics["retry_sleep_seconds"],
        "attempt_count": retry_metrics["attempt_count"],
        "input_tokens": response.usage.input_tokens,
        "output_tokens": response.usage.output_tokens,
        "model_version": getattr(response, "model", kwargs["model"]),
    }


def call_gpt(row: dict, condition: str, pdf_path: str | Path = None, model_name: str = "gpt") -> dict:
    uses_pdf = condition in {"pdf", "pdf_websearch", "pdf_structure"}
    uses_websearch = condition in {"websearch", "pdf_websearch"}

    if uses_pdf and pdf_path:
        input_value = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_file",
                        "filename": Path(pdf_path).name,
                        "file_data": f"data:application/pdf;base64,{get_pdf_base64(pdf_path)}",
                    },
                    {
                        "type": "input_text",
                        "text": build_user_prompt(row),
                    },
                ],
            }
        ]
    else:
        input_value = build_user_prompt(row)

    payload = {
        "model": resolve_model_name(model_name, condition),
        "instructions": get_system_prompt(row, condition),
        "input": input_value,
        "max_output_tokens": MAX_RESPONSE_TOKENS,
    }
    if uses_websearch:
        payload["tools"] = [{"type": "web_search"}]

    response, retry_metrics = with_retries(
        f"{resolve_model_name(model_name)}|{condition}",
        lambda: call_openai_responses_api(payload),
    )

    answer = response["raw"].strip()
    parsed_output = parse_model_output(answer)

    return {
        "raw": parsed_output["normalized_raw"],
        "parsed": parsed_output["decision"],
        "reason": parsed_output["reason"],
        "search_queries": " | ".join(response.get("search_queries", [])),
        "latency": retry_metrics["api_latency"],
        "total_latency": retry_metrics["total_latency"],
        "retry_sleep_seconds": retry_metrics["retry_sleep_seconds"],
        "attempt_count": retry_metrics["attempt_count"],
        "input_tokens": response["input_tokens"],
        "output_tokens": response["output_tokens"],
        "model_version": response["model_version"],
    }


def call_gemini(row: dict, condition: str, pdf_path: str | Path = None, model_name: str = "gemini") -> dict:
    require_api_key("GEMINI", GEMINI_API_KEY)
    uses_pdf_native = condition in {"pdf", "pdf_websearch", "pdf_structure"}
    uses_text = condition in {"text_md", "text_md_websearch"}
    uses_websearch = condition in {"websearch", "pdf_websearch", "text_md_websearch"}

    try:
        from google import genai
        from google.genai import types

        gemini_http_options = types.HttpOptions(
            timeout=GEMINI_HTTP_TIMEOUT_MS,
            retryOptions=types.HttpRetryOptions(attempts=1),
        )
        client = genai.Client(
            api_key=GEMINI_API_KEY,
            http_options=gemini_http_options,
        )
        config_kwargs = {
            "systemInstruction": get_system_prompt(row, condition),
            "maxOutputTokens": MAX_RESPONSE_TOKENS,
            "httpOptions": gemini_http_options,
        }
        if uses_websearch:
            config_kwargs["tools"] = [types.Tool(googleSearch=types.GoogleSearch())]

        parts = []
        if uses_pdf_native and pdf_path:
            parts.append(
                types.Part.from_bytes(
                    data=get_pdf_bytes(pdf_path),
                    mime_type="application/pdf",
                )
            )
        elif uses_text and pdf_path:
            parts.append(types.Part.from_text(text=f"[급여 고시 원문]\n{get_pdf_text(pdf_path)}"))
        parts.append(types.Part.from_text(text=build_user_prompt(row)))

        response, retry_metrics = with_retries(
            f"{resolve_model_name(model_name)}|{condition}",
            lambda: client.models.generate_content(
                model=resolve_model_name(model_name, condition),
                contents=parts,
                config=types.GenerateContentConfig(**config_kwargs),
            ),
        )

        answer = (getattr(response, "text", "") or "").strip()
        usage = getattr(response, "usage_metadata", None)

        # Extract search queries
        search_queries: list[str] = []
        try:
            candidate = response.candidates[0]
            grounding_meta = getattr(candidate, "grounding_metadata", None)
            if grounding_meta:
                search_queries = list(grounding_meta.web_search_queries or [])
        except (IndexError, AttributeError):
            pass

        parsed_output = parse_model_output(answer)

        return {
            "raw": parsed_output["normalized_raw"],
            "parsed": parsed_output["decision"],
            "reason": parsed_output["reason"],
            "search_queries": " | ".join(search_queries),
            "latency": retry_metrics["api_latency"],
            "total_latency": retry_metrics["total_latency"],
            "retry_sleep_seconds": retry_metrics["retry_sleep_seconds"],
            "attempt_count": retry_metrics["attempt_count"],
            "input_tokens": getattr(usage, "prompt_token_count", 0) if usage else 0,
            "output_tokens": getattr(usage, "candidates_token_count", 0) if usage else 0,
            "model_version": resolve_model_name(model_name, condition),
        }
    except ImportError:
        import google.generativeai as genai

        genai.configure(api_key=GEMINI_API_KEY)
        tools = [{"google_search": {}}] if uses_websearch else None
        model = genai.GenerativeModel(
            model_name=resolve_model_name(model_name, condition),
            system_instruction=get_system_prompt(row, condition),
            tools=tools,
        )

        parts = []
        if uses_pdf_native and pdf_path:
            parts.append({
                "inline_data": {
                    "mime_type": "application/pdf",
                    "data": get_pdf_base64(pdf_path),
                }
            })
        elif uses_text and pdf_path:
            parts.append({"text": f"[급여 고시 원문]\n{get_pdf_text(pdf_path)}"})
        parts.append({"text": build_user_prompt(row)})

        response, retry_metrics = with_retries(
            f"{resolve_model_name(model_name)}|{condition}",
            lambda: model.generate_content(parts),
        )

        answer = response.text.strip()

        # Extract search queries (legacy SDK)
        search_queries_legacy: list[str] = []
        try:
            grounding_meta = response.candidates[0].grounding_metadata
            search_queries_legacy = list(grounding_meta.web_search_queries or [])
        except (IndexError, AttributeError):
            pass

        parsed_output = parse_model_output(answer)

        return {
            "raw": parsed_output["normalized_raw"],
            "parsed": parsed_output["decision"],
            "reason": parsed_output["reason"],
            "search_queries": " | ".join(search_queries_legacy),
            "latency": retry_metrics["api_latency"],
            "total_latency": retry_metrics["total_latency"],
            "retry_sleep_seconds": retry_metrics["retry_sleep_seconds"],
            "attempt_count": retry_metrics["attempt_count"],
            "input_tokens": response.usage_metadata.prompt_token_count,
            "output_tokens": response.usage_metadata.candidates_token_count,
            "model_version": resolve_model_name(model_name, condition),
        }


# ──────────────────────────────────────────────
# 5. Response parsing
# ──────────────────────────────────────────────

def parse_answer(raw: str) -> str:
    """Normalize a model response to eligible / ineligible / undeterminable / invalid"""
    raw = raw.lower().strip()

    # direct matching
    if raw in ("eligible", "ineligible", "undeterminable"):
        return raw

    if raw.startswith("ineligib"):
        return "ineligible"
    if raw.startswith("eligib"):
        return "eligible"
    if raw.startswith("undetermin"):
        return "undeterminable"

    if "not eligible" in raw:
        return "ineligible"

    # Korean output mapping
    if any(k in raw for k in ["부적격", "ineligible"]):
        return "ineligible"
    if any(k in raw for k in ["적격", "eligible"]):
        return "eligible"
    if any(k in raw for k in ["판정불가", "알 수 없", "undeterminable", "undetermined"]):
        return "undeterminable"

    return "invalid"  # parse failure


# ──────────────────────────────────────────────
# 6. Main evaluation loop
# ──────────────────────────────────────────────

MODEL_CALLERS = {
    "claude": call_claude,
    "gpt": call_gpt,
    "gemini": call_gemini,
}

CONDITIONS = ["websearch", "pdf", "pdf_websearch", "text_md", "text_md_websearch", "pdf_structure"]
PDF_REQUIRED_CONDITIONS = {"pdf", "pdf_websearch", "text_md", "text_md_websearch", "pdf_structure"}


def run_evaluation(
    df: pd.DataFrame,
    models: list = None,
    conditions: list = None,
    dry_run: bool = False,
    delay: float = 0.5,
    task_plan: pd.DataFrame | None = None,
    incremental_csv_path: str | Path | None = None,
    n_runs: int = 1,
) -> pd.DataFrame:
    """
    Run the full evaluation.
    dry_run=True: validate structure without real API calls
    """
    models = models or ["claude", "gpt", "gemini"]
    conditions = conditions or CONDITIONS

    for cancer, path in PDF_PATHS.items():
        if Path(path).exists():
            print(f"PDF check: {cancer} -> {Path(path).name}")
        else:
            raise FileNotFoundError(f"PDF file not found: {path}")

    if task_plan is None:
        task_records = []
        for run_id in range(n_runs):
            for _, row in df.iterrows():
                for condition in conditions:
                    for model_name in models:
                        task_records.append({
                            "row": row,
                            "model_name": model_name,
                            "condition": condition,
                            "run_id": run_id,
                        })
    else:
        task_records = [
            {
                "row": task,
                "model_name": task["__model"],
                "condition": task["__condition"],
                "run_id": task.get("__run_id", 0),
            }
            for _, task in task_plan.iterrows()
        ]

    results = []
    total = len(task_records)
    count = 0

    for task in task_records:
        row = task["row"]
        cancer = row["cancer_type"]
        pdf_path = PDF_PATHS.get(cancer)
        model_name = task["model_name"]
        condition = task["condition"]
        run_id = task.get("run_id", 0)

        count += 1
        print(f"[{count}/{total}] {model_name} | {condition} | {row['ID']}", end=" ")

        if dry_run:
            dry_predicted = (
                "eligible"
            )
            result = {
                "raw": json.dumps({"decision": "eligible", "reason": "Dry run placeholder."}),
                "parsed": dry_predicted,
                "reason": "Dry run placeholder.",
                "latency": 0.0,
                "total_latency": 0.0,
                "retry_sleep_seconds": 0.0,
                "attempt_count": 0,
                "input_tokens": 0,
                "output_tokens": 0,
            }
        else:
            try:
                provider = resolve_model_provider(model_name)
                fn = MODEL_CALLERS[provider]
                selected_pdf_path = pdf_path if condition in PDF_REQUIRED_CONDITIONS else None
                result = fn(row, condition, selected_pdf_path, model_name=model_name)

                time.sleep(delay)
            except Exception as e:
                print(f"ERROR: {e}")
                result = {
                    "raw": str(e),
                    "parsed": "error",
                    "reason": "",
                    "latency": float(getattr(e, "_retry_metrics", {}).get("api_latency", 0.0)),
                    "total_latency": float(getattr(e, "_retry_metrics", {}).get("total_latency", 0.0)),
                    "retry_sleep_seconds": float(getattr(e, "_retry_metrics", {}).get("retry_sleep_seconds", 0.0)),
                    "attempt_count": int(getattr(e, "_retry_metrics", {}).get("attempt_count", 0)),
                    "input_tokens": 0,
                    "output_tokens": 0,
                }

        correct = result["parsed"] == row["expected"]
        print(f"→ {result['parsed']} ({'✓' if correct else '✗'})")

        result_row = {
            "id": row["ID"],
            "cancer_type": cancer,
            "class": row["class"],
            "regimen_code": row.get("regimen_code", ""),
            "regimen": row.get("regimen", ""),
            "expected": row["expected"],
            "model": model_name,
            "model_provider": resolve_model_provider(model_name),
            "model_version": result.get("model_version", resolve_model_name(model_name, condition)),
            "generation_settings_policy": GENERATION_SETTINGS_POLICY,
            "condition": condition,
            "run_id": run_id,
            "predicted": result["parsed"],
            "reason": result.get("reason", ""),
            "correct": correct,
            "raw_response": result["raw"],
            "search_queries": result.get("search_queries", ""),
            "latency": result["latency"],
            "total_latency": result.get("total_latency", result["latency"]),
            "retry_sleep_seconds": result.get("retry_sleep_seconds", 0.0),
            "attempt_count": result.get("attempt_count", 1 if not dry_run else 0),
            "input_tokens": result["input_tokens"],
            "output_tokens": result["output_tokens"],
            "timestamp": datetime.now().isoformat(),
        }
        results.append(result_row)
        if incremental_csv_path is not None:
            append_result_csv_row(incremental_csv_path, result_row)

    return pd.DataFrame(results)


# ──────────────────────────────────────────────
# 7. Entry point
# ──────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="K-NHIB Benchmark Evaluation")
    parser.add_argument(
        "--models",
        nargs="+",
        default=["claude", "gpt", "gemini"],
        choices=SUPPORTED_MODEL_NAMES,
        help="Model alias or actual model code to evaluate",
    )
    parser.add_argument(
        "--conditions",
        nargs="+",
        default=["websearch", "pdf", "pdf_websearch"],
        choices=CONDITIONS,
        help="Evaluation conditions (websearch pdf pdf_websearch)",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate structure without API calls")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Delay between calls (seconds)")
    parser.add_argument("--cancer", nargs="+",
                        choices=["자궁경부암", "자궁내막암", "난소암"],
                        help="Run only a specific cancer type")
    parser.add_argument(
        "--retry-errors-from",
        help="Re-run only predicted=error items from an existing results csv/xlsx",
    )
    parser.add_argument(
        "--n-runs", type=int, default=1,
        help="Number of repeated runs (default 1). 3-5 recommended for mean/SD/CI estimation",
    )
    args = parser.parse_args()

    # Load data
    df = load_benchmark(EXCEL_PATH)

    # Cancer filter
    if args.cancer:
        df = df[df["cancer_type"].isin(args.cancer)]
        print(f"Filter applied: {args.cancer} -> {len(df)} cases")

    validate_runtime(args.models, args.dry_run)

    previous_results_df = None
    retry_task_plan = None
    if args.retry_errors_from:
        previous_results_df = load_saved_results(args.retry_errors_from)
        retry_task_plan = build_retry_task_plan(
            benchmark_df=df,
            previous_results=previous_results_df,
            models=args.models,
            conditions=args.conditions,
        )
        print(
            "Re-run targets: "
            f"{len(retry_task_plan)} failed combinations "
            f"(source={Path(args.retry_errors_from).name})"
        )
        if retry_task_plan.empty:
            print("No error items to re-run.")
            raise SystemExit(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_label = "knhib_results_retry" if previous_results_df is not None else "knhib_results"
    out_path = OUTPUT_DIR / f"{run_label}_{timestamp}.csv"
    print(f"Intermediate save: {out_path}")

    # Run evaluation
    rerun_df = run_evaluation(
        df,
        models=args.models,
        conditions=args.conditions,
        dry_run=args.dry_run,
        delay=args.delay,
        task_plan=retry_task_plan,
        incremental_csv_path=out_path,
        n_runs=args.n_runs,
    )

    results_df = (
        merge_results_frames(previous_results_df, rerun_df)
        if previous_results_df is not None
        else rerun_df
    )

    # Save
    results_df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"\nResults saved: {out_path}")

    # Also save raw results as Excel. Scoring/metrics are produced by aggregate_results.py.
    if not args.dry_run:
        excel_path = OUTPUT_DIR / f"{run_label}_{timestamp}.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            results_df.to_excel(writer, sheet_name="raw", index=False)
        print(f"Excel saved: {excel_path}")
